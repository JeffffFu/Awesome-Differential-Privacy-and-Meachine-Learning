from data.data_load.data_load_Mnist import dataload_mnist_60000
from data.get_data import get_data
from data.util.sampling import get_data_loaders_uniform_without_replace
from model.CNN import CNN_tanh, CNN
from optimizer.dp_optimizer import DPSGD, DPAdam
import torch
from torch import nn
from torch.utils.data import TensorDataset
from torch.utils.data import DataLoader

from privacy_analysis.RDP.compute_rdp import compute_rdp
from privacy_analysis.RDP.rdp_convert_dp import compute_eps
from train_and_validation.train_with_dp import train_dynamic_add_noise
from train_and_validation.validation import validation
from openpyxl import Workbook
import time
import copy
from scipy.stats import bernoulli
import numpy as np


def centralization_train_dynamic_add_noise_SimulatedAnnealing(train_data, test_data, model,batch_size, numEpoch, learning_rate,momentum,delta,max_norm,sigma,k,mu_0):


    optimizer = DPAdam(
        l2_norm_clip=max_norm,      #裁剪范数
        noise_multiplier=sigma,
        minibatch_size=batch_size,        #几个样本梯度进行一次梯度下降
        microbatch_size=1,                #几个样本梯度进行一次裁剪，这里选择逐样本裁剪
        params=model.parameters(),
        lr=learning_rate,
        # momentum=momentum
    )
    centralized_criterion = nn.CrossEntropyLoss()

    # 包装抽样函数
    minibatch_size = batch_size  # 这里比较好的取值是根号n，n为每个客户端的样本数
    microbatch_size = 1  #这里默认1就好
    iterations = 1  # n个batch，这边就定一个，每次训练采样一个Lot
    minibatch_loader, microbatch_loader = get_data_loaders_uniform_without_replace(minibatch_size, microbatch_size, iterations)     #无放回均匀采样

    # Dataset是一个包装类，用来将数据包装为Dataset类，然后传入DataLoader中，我们再使用DataLoader这个类来更加快捷的对数据进行操作。
    # DataLoader是一个比较重要的类，它为我们提供的常用操作有：batch_size(每个batch的大小), shuffle(是否进行shuffle操作)
    # DataLoader里的数据要通过for i ,data in enumerate(DataLoader)才能打印出来

    #train数据在下面进行采样，这边不做dataloader
    test_dl = torch.utils.data.DataLoader(
        test_data, batch_size=batch_size, shuffle=False)

    print("------ Centralized Model ------")
    rdp=0
    orders = (list(range(2, 64)) + [128, 256, 512])  # 默认的lamda
    epsilon_list=[]
    result_loss_list=[]
    result_acc_list=[]
    last_central_test_loss=100000.0
    last_centralized_model=model
    t=0

    for epoch in range(numEpoch):


        train_dl = minibatch_loader(train_data)     #抽样

        central_train_loss, central_train_accuracy = train_dynamic_add_noise(model, train_dl, centralized_criterion, optimizer)

        central_test_loss, central_test_accuracy = validation(model, test_dl)


        if central_test_loss-last_central_test_loss>0:
            p = np.exp(-(central_test_loss - last_central_test_loss) * k * t)  # 概率p随着epoch增大而减小，随着损失值的差增大而减小，看出来是接收的概率
            print("central_test_loss {:7.4f}".format(central_test_loss) + "|  last_central_test_loss {:7.4f}".format(
                last_central_test_loss) + "|  接受新模型的概率p:", p)
            rv = bernoulli.rvs(p)  # 伯努利分布，p的概率输出1，1-p的概率输出0。输出1表示接受新模型

            if rv==0 and mu<mu_0 :      #==1接受模型，那么==0就是不接受模型更新，返回到上一次
                print("不接受模型更新,mu:",mu)
                #把上一轮模型保存的参数传回来
                model.load_state_dict(last_centralized_model.state_dict(),strict=True)
                mu=mu+1
            else:

                last_central_test_loss = central_test_loss
                last_central_test_acc = central_test_accuracy

                #copy.deepcopy可以只拷贝model里面的值，而不会随着model变化而改变
                last_centralized_model = copy.deepcopy(model)
                t=t+1
                print("接受新模型，当前接受的迭代轮次t：", format(t))
                mu=0

        else:
            last_central_test_loss = central_test_loss
            last_central_test_acc = central_test_accuracy
            last_centralized_model = copy.deepcopy(model)
            t = t + 1
            mu = 0
            print("接受新模型，当前接受的迭代轮次t：", format(t))


        # 这里要每次根据simga累加它的RDP，循环结束再转为eps，这里的epoch系数直接设为t,表示成功迭代的次数
        rdp_every_epoch=compute_rdp(batch_size/len(train_data), sigma, t, orders)
        rdp=rdp+rdp_every_epoch
        epsilon, best_alpha = compute_eps(orders, rdp, delta)
        epsilon_list.append(epsilon)

        result_loss_list.append(central_test_loss)
        result_acc_list.append(central_test_accuracy)


        print("epoch: {:3.0f}".format(epoch + 1) + " | epsilon: {:7.4f}".format(
        epsilon) + " | best_alpha: {:7.4f}".format(best_alpha)  )

        if (epsilon > 3):
            wb = Workbook()
            sheet = wb.active
            sheet.title = "result"
            sheet.cell(1, 3).value = "acc"
            sheet.cell(1, 4).value = "eps"
            sheet.cell(1, 5).value = "sigma"
            # sheet.cell(1, 6).value="实验时间：{}".format(datetime.datetime.now())
            sheet.cell(1, 7).value = "| batch_size:{}".format(batch_size) + "| learning_rate:{}".format(
                learning_rate) + "| sigma:{}".format(sigma) + "| max_norm:{}".format(max_norm) + "| numepoch:{}".format(
                numEpoch)
            # sheet.cell(1, 8).value = "mnist数据自适应范数裁剪，不是逐层"
            for i in range(len(result_loss_list)):
                sheet.cell(i + 2, 2).value = result_loss_list[i]
                sheet.cell(i + 2, 3).value = result_acc_list[i]
                sheet.cell(i + 2, 4).value = epsilon_list[i]
            wb.save("../result/{}.xlsx".format(int(time.time())))
            break

    print("------ Training finished ------")

if __name__=="__main__":

    train_data, test_data = get_data('mnist', augment=False)
    model = CNN()
    batch_size = 256
    learning_rate = 0.002
    numEpoch = 1500
    sigma = 1.1
    momentum = 0.9
    delta = 10 ** (-5)
    max_norm = 0.1
    k=10  #这个控制模拟退退火下接受的概率
    mu_0=5  #表示连续n次不接受新解后强制接受
    centralization_train_dynamic_add_noise_SimulatedAnnealing(train_data, test_data, model,batch_size, numEpoch, learning_rate,momentum,delta,max_norm,sigma,k,mu_0)